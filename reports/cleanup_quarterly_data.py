"""
Pre-report data cleanup for AQUILA Quarterly Reports (Office & Industrial).

Runs before generate_office_report.py / generate_industrial_report.py to:
  1. Standardize street abbreviations and cardinal directions in address/name columns
     across all quarterly Excel and CSV files.
  2. Ensure the Major Leases and Sales file has a 'Vertical Format' tab derived
     from the 'Major Sales' sheet (creates it if missing).
  3. Sort Major Leases by SF ascending, format SF with commas, and match building
     names to Supabase report_name.
  4. Consolidate portfolio sales (same buyer + seller) in Major Sales.
  5. Verify Under Construction buildings exist in Supabase inventory (warnings only).
  6. Sort Planned/Proposed by Submarket (A-Z), then SF (ascending).

Usage:
    python reports/cleanup_quarterly_data.py          # uses report_config paths
    python reports/cleanup_quarterly_data.py --dry-run  # preview only, no writes
"""

import os
import re
import sys
import argparse

import pandas as pd
import openpyxl

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from reports import report_config as config


# ---------------------------------------------------------------------------
# Report type detection
# ---------------------------------------------------------------------------

def _detect_report_type(cfg):
    """Detect whether this is an office or industrial config.
    Returns 'office' or 'industrial'.
    """
    if hasattr(cfg, 'PROPERTY_TYPES'):
        return 'industrial'
    return 'office'


# ---------------------------------------------------------------------------
# Abbreviation standardization
# ---------------------------------------------------------------------------

# Patterns applied in order. Each key is a regex pattern; value is replacement.
# The regex uses \b word boundaries so we don't clobber partial words.
# Replacements that already end with '.' are left untouched (the lookahead below
# ensures we only replace when NOT already followed by '.').
ABBREVIATIONS = {
    r'\bDr\b':   'Dr.',
    r'\bBlvd\b': 'Blvd.',
    r'\bLn\b':   'Ln.',
    r'\bRd\b':   'Rd.',
    r'\bSt\b':   'St.',
    r'\bCv\b':   'Cv.',
    r'\bAve\b':  'Ave.',
    r'\bPkwy\b': 'Pkwy.',
    r'\bCir\b':  'Cir.',
    r'\bCt\b':   'Ct.',
    r'\bPl\b':   'Pl.',
    r'\bTer\b':  'Ter.',
    r'\bWy\b':   'Wy.',
    r'\bTrl\b':  'Trl.',
    r'\bHwy\b':  'Hwy.',
    r'\bSq\b':   'Sq.',
    r'\bLp\b':   'Lp.',
    # Cardinal directions — negative lookbehind (?<!\.) prevents expanding
    # a letter that already follows a period (e.g. "Bldg. E" stays "Bldg. E")
    r'(?<!\.)\bN\b':    'N.',
    r'(?<!\.)\bS\b':    'S.',
    r'(?<!\.)\bE\b':    'E.',
    r'(?<!\.)\bW\b':    'W.',
    r'\bNE\b':   'N.E.',
    r'\bNW\b':   'N.W.',
    r'\bSE\b':   'S.E.',
    r'\bSW\b':   'S.W.',
}

# Column name patterns that should be standardized (case-insensitive)
ADDRESS_COLUMN_PATTERNS = [
    r'property[\s_]*name',
    r'property[\s_]*address',
    r'building[\s_]*park',
    r'building[\s_]*name',
    r'tenant',
]


def standardize_abbreviations(text):
    """Add periods to common street abbreviations and cardinal directions."""
    if pd.isna(text) or not isinstance(text, str):
        return text
    result = text
    for pattern, replacement in ABBREVIATIONS.items():
        # Only replace when NOT already followed by a period
        result = re.sub(f'{pattern}(?!\\.)', replacement, result, flags=re.IGNORECASE)
    return result


def find_address_columns(columns):
    """Return list of column names matching address/name patterns."""
    matching = []
    for col in columns:
        col_norm = str(col).lower().replace(' ', '_')
        for pat in ADDRESS_COLUMN_PATTERNS:
            if re.search(pat, col_norm):
                matching.append(col)
                break
    return matching


def _apply_to_worksheet(ws, dry_run=False):
    """
    Apply abbreviation standardization to visible address columns in an openpyxl worksheet.
    Returns count of cells changed.
    """
    # Read header row (row 1) to find address columns
    header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    address_col_indices = []  # 1-based column indices
    for i, col_name in enumerate(header, start=1):
        if col_name is None:
            continue
        col_norm = str(col_name).lower().replace(' ', '_')
        for pat in ADDRESS_COLUMN_PATTERNS:
            if re.search(pat, col_norm):
                address_col_indices.append(i)
                break

    changed = 0
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in address_col_indices:
            cell = ws.cell(row=row_idx, column=col_idx)
            original = cell.value
            updated = standardize_abbreviations(original)
            if updated != original:
                if not dry_run:
                    cell.value = updated
                changed += 1
    return changed


def process_excel_file(path, dry_run=False):
    """Standardize address abbreviations across all visible sheets of an Excel file."""
    wb = openpyxl.load_workbook(path, data_only=False)
    visible_sheets = [ws for ws in wb.worksheets if ws.sheet_state == 'visible']
    total_changed = 0

    for ws in visible_sheets:
        n = _apply_to_worksheet(ws, dry_run=dry_run)
        if n:
            print(f"    Sheet '{ws.title}': {n} cell(s) updated")
            total_changed += n

    if total_changed and not dry_run:
        wb.save(path)
    return total_changed


def process_csv_file(path, dry_run=False):
    """Standardize address abbreviations in address columns of a CSV file."""
    df = pd.read_csv(path)
    cols = find_address_columns(df.columns)
    total_changed = 0

    for col in cols:
        original = df[col].copy()
        df[col] = df[col].apply(standardize_abbreviations)
        n = (df[col] != original).sum()
        if n:
            print(f"    Column '{col}': {n} cell(s) updated")
            total_changed += n

    if total_changed and not dry_run:
        df.to_csv(path, index=False)
    return total_changed


# ---------------------------------------------------------------------------
# Vertical Format tab for Major Sales
# ---------------------------------------------------------------------------

VERTICAL_SHEET_NAME = 'Vertical Format'


def _ensure_vertical_format(path, dry_run=False):
    """
    Check if the Major Leases and Sales workbook already has a 'Vertical Format'
    sheet. If not, create one from the 'Major Sales' sheet with a vertical key-value
    layout (sorted by square footage descending).
    """
    wb = openpyxl.load_workbook(path, data_only=False)

    if VERTICAL_SHEET_NAME in wb.sheetnames:
        print(f"    '{VERTICAL_SHEET_NAME}' tab already exists -- skipping.")
        return False

    if 'Major Sales' not in wb.sheetnames:
        print(f"    'Major Sales' sheet not found in {os.path.basename(path)} -- skipping vertical format.")
        return False

    print(f"    Creating '{VERTICAL_SHEET_NAME}' tab from 'Major Sales'...")

    # Read Major Sales into a DataFrame
    df = pd.read_excel(path, sheet_name='Major Sales')

    # Sort by square footage descending (try common column names)
    for sf_col in ('Square Feet', 'Size', 'SF', 'GLA'):
        if sf_col in df.columns:
            df = df.sort_values(sf_col, ascending=False)
            df[sf_col] = df[sf_col].apply(
                lambda x: f"{int(x):,}" if pd.notna(x) and str(x).replace('.', '', 1).isdigit()
                else x
            )
            break

    # Build vertical key-value rows
    vertical_rows = []
    for i, (_, row) in enumerate(df.iterrows()):
        if i > 0:
            vertical_rows.append(('', ''))  # blank separator between records
        for col_name, value in row.items():
            vertical_rows.append((col_name, value))

    if dry_run:
        print(f"    [dry-run] Would add {len(df)} records ({len(vertical_rows)} rows) to '{VERTICAL_SHEET_NAME}'.")
        return False

    # Add the new sheet
    ws_vert = wb.create_sheet(VERTICAL_SHEET_NAME)
    ws_vert.append(['Attribute', 'Value'])  # header
    for row in vertical_rows:
        ws_vert.append(list(row))

    wb.save(path)
    print(f"    Added '{VERTICAL_SHEET_NAME}' with {len(df)} sale records.")
    return True


# ---------------------------------------------------------------------------
# Supabase inventory loading & name matching
# ---------------------------------------------------------------------------

def _load_inventory(report_type):
    """
    Load the inventory table from Supabase for building name matching.
    Returns a DataFrame with columns: costar_property_id, property_name,
    building_park, report_name, aquila_competitive_set, property_address.
    Returns empty DataFrame if Supabase is unavailable.
    """
    try:
        from aquila.connectors.supabase import get_supabase_client
        supabase = get_supabase_client(use_service_role=True)
        table_name = 'inventory_office' if report_type == 'office' else 'inventory_industrial'
        response = supabase.table(table_name).select(
            'costar_property_id, property_name, building_park, report_name, '
            'aquila_competitive_set, property_address'
        ).execute()
        df = pd.DataFrame(response.data)
        print(f"    Loaded {len(df)} rows from {table_name}")
        return df
    except Exception as e:
        print(f"    Warning: Could not load inventory from Supabase ({e})")
        return pd.DataFrame()


def _normalize_name(name):
    """Normalize a building name for comparison.
    Lowercase, strip periods/commas/dashes, collapse whitespace.
    """
    if pd.isna(name) or not isinstance(name, str):
        return ''
    result = re.sub(r'[.,\-\'\"]+', '', name.lower())
    result = re.sub(r'\s+', ' ', result).strip()
    return result


def _build_inventory_lookup(inventory_df):
    """
    Build a dict for fast normalized name -> report_name lookup.
    Matches on both property_name and building_park.
    Only entries with non-empty report_name are included.
    Returns: dict[normalized_name] -> report_name
    """
    lookup = {}
    if inventory_df.empty:
        return lookup

    for _, row in inventory_df.iterrows():
        report_name = row.get('report_name')
        if pd.isna(report_name) or not str(report_name).strip():
            continue
        report_name = str(report_name).strip()

        pname = _normalize_name(row.get('property_name'))
        if pname:
            lookup[pname] = report_name

        bpark = _normalize_name(row.get('building_park'))
        if bpark and bpark != pname:
            lookup[bpark] = report_name

    return lookup


# ---------------------------------------------------------------------------
# Helper: rewrite a DataFrame back to a specific sheet in an Excel workbook
# ---------------------------------------------------------------------------

def _rewrite_sheet(path, sheet_name, df):
    """Delete and recreate a sheet in an Excel workbook, preserving other sheets.
    Inserts the new sheet at the same index position as the original.
    """
    wb = openpyxl.load_workbook(path, data_only=False)
    if sheet_name in wb.sheetnames:
        sheet_idx = wb.sheetnames.index(sheet_name)
        del wb[sheet_name]
    else:
        sheet_idx = len(wb.sheetnames)

    ws = wb.create_sheet(sheet_name, sheet_idx)
    # Write header
    for col_idx, col_name in enumerate(df.columns, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)
    # Write data
    for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
        for col_idx, val in enumerate(row, start=1):
            # Convert numpy types to native Python for openpyxl
            if pd.notna(val):
                ws.cell(row=row_idx, column=col_idx, value=val)
            else:
                ws.cell(row=row_idx, column=col_idx, value=None)
    wb.save(path)


# ---------------------------------------------------------------------------
# Helper: detect SF column from common name variants
# ---------------------------------------------------------------------------

def _find_sf_column(columns):
    """Return the first matching SF column name, or None."""
    for candidate in ('SF Leased', 'Size (SF)', 'Size', 'Square Feet', 'SF', 'GLA'):
        if candidate in columns:
            return candidate
    return None


def _find_column_by_keyword(columns, keyword):
    """Return first column whose name contains keyword (case-insensitive), or None."""
    for col in columns:
        if keyword.lower() in str(col).lower():
            return col
    return None


# ---------------------------------------------------------------------------
# Major Leases cleanup
# ---------------------------------------------------------------------------

def _cleanup_major_leases(path, inventory_lookup, dry_run=False):
    """
    Clean up the Major Leases sheet:
    1. Sort by SF ascending (smallest to largest)
    2. Add commas to SF column if missing
    3. Match building names against Supabase inventory; replace with report_name
    Returns count of changes made.
    """
    try:
        xl = pd.ExcelFile(path)
        if 'Major Leases' not in xl.sheet_names:
            print(f"    'Major Leases' sheet not found -- skipping lease cleanup")
            return 0
    except Exception:
        return 0

    df = pd.read_excel(path, sheet_name='Major Leases')
    if df.empty:
        return 0

    changed = 0

    # --- Identify SF column ---
    sf_col = _find_sf_column(df.columns)
    if sf_col is None:
        print(f"    Warning: No SF column found in Major Leases")
        return 0

    # Clean and convert SF to numeric for sorting
    df['_sf_numeric'] = pd.to_numeric(
        df[sf_col].astype(str).str.replace(',', '', regex=False),
        errors='coerce'
    )

    # --- Sort by SF ascending ---
    df_sorted = df.sort_values('_sf_numeric', ascending=True, na_position='last')
    if not df_sorted['_sf_numeric'].equals(df['_sf_numeric']):
        changed += 1
        print(f"    Sorted Major Leases by {sf_col} ascending")
    df = df_sorted.reset_index(drop=True)

    # --- Format SF with commas ---
    sf_formatted = df['_sf_numeric'].apply(
        lambda x: f"{int(x):,}" if pd.notna(x) else ''
    )
    original_sf = df[sf_col].astype(str)
    sf_changes = (sf_formatted != original_sf).sum()
    if sf_changes > 0:
        df[sf_col] = sf_formatted
        changed += sf_changes
        print(f"    Formatted {sf_changes} SF value(s) with commas")

    # Drop helper column
    df = df.drop(columns=['_sf_numeric'])

    # --- Building name matching against inventory ---
    building_col = _find_column_by_keyword(df.columns, 'property name')
    if building_col is None:
        building_col = _find_column_by_keyword(df.columns, 'building')

    if building_col and inventory_lookup:
        name_changes = 0
        for idx, row in df.iterrows():
            original_name = row[building_col]
            if pd.isna(original_name):
                continue
            normalized = _normalize_name(original_name)
            if normalized in inventory_lookup:
                report_name = inventory_lookup[normalized]
                if str(original_name).strip() != report_name:
                    df.at[idx, building_col] = report_name
                    name_changes += 1
        if name_changes:
            changed += name_changes
            print(f"    Matched {name_changes} building name(s) to Supabase report_name")

    # --- Write back ---
    if changed > 0 and not dry_run:
        _rewrite_sheet(path, 'Major Leases', df)
    elif dry_run and changed > 0:
        print(f"    [dry-run] Would make {changed} change(s) to Major Leases")

    return changed


# ---------------------------------------------------------------------------
# Major Sales portfolio consolidation
# ---------------------------------------------------------------------------

def _cleanup_major_sales(path, dry_run=False):
    """
    Consolidate portfolio sales in Major Sales sheet:
    - Merge rows with same buyer AND same seller (regardless of submarket)
    - Combine SF values (sum) and join building names with commas
    - Keep submarket from first row; append others if different
    Returns count of rows consolidated.
    """
    try:
        xl = pd.ExcelFile(path)
        if 'Major Sales' not in xl.sheet_names:
            print(f"    'Major Sales' sheet not found -- skipping sales consolidation")
            return 0
    except Exception:
        return 0

    df = pd.read_excel(path, sheet_name='Major Sales')
    if df.empty or len(df) < 2:
        return 0

    # Identify buyer/seller columns
    buyer_col = _find_column_by_keyword(df.columns, 'buyer')
    seller_col = _find_column_by_keyword(df.columns, 'seller')

    if not buyer_col or not seller_col:
        print(f"    Warning: Could not identify Buyer/Seller columns in Major Sales")
        return 0

    # Identify other key columns
    name_col = _find_column_by_keyword(df.columns, 'property name')
    sf_col = _find_sf_column(df.columns)
    sub_col = _find_column_by_keyword(df.columns, 'submarket')
    if sub_col is None:
        sub_col = _find_column_by_keyword(df.columns, 'market')

    if not sf_col:
        print(f"    Warning: No SF column found in Major Sales")
        return 0

    # Clean SF to numeric
    df['_sf_numeric'] = pd.to_numeric(
        df[sf_col].astype(str).str.replace(',', '', regex=False),
        errors='coerce'
    )

    # Group by (buyer, seller) normalized
    df['_buyer_norm'] = df[buyer_col].astype(str).str.strip().str.lower()
    df['_seller_norm'] = df[seller_col].astype(str).str.strip().str.lower()

    groups = df.groupby(['_buyer_norm', '_seller_norm'], sort=False)
    if len(groups) == len(df):
        # No consolidation needed
        return 0

    consolidated_rows = []
    consolidation_count = 0

    for (_buyer, _seller), group in groups:
        if len(group) == 1:
            row_dict = group.iloc[0].to_dict()
            consolidated_rows.append(row_dict)
        else:
            consolidation_count += len(group) - 1
            base_row = group.iloc[0].to_dict()

            # Combine building names
            if name_col:
                names = group[name_col].dropna().astype(str).unique().tolist()
                base_row[name_col] = ', '.join(names)

            # Sum SF
            base_row['_sf_numeric'] = group['_sf_numeric'].sum()

            # Combine submarkets if different
            if sub_col:
                subs = group[sub_col].dropna().astype(str).unique().tolist()
                if len(subs) > 1:
                    base_row[sub_col] = ', '.join(subs)

            consolidated_rows.append(base_row)

    if consolidation_count == 0:
        return 0

    new_df = pd.DataFrame(consolidated_rows)

    # Format SF with commas from the numeric column
    new_df[sf_col] = new_df['_sf_numeric'].apply(
        lambda x: f"{int(x):,}" if pd.notna(x) else ''
    )

    # Drop helper columns
    for col in ['_buyer_norm', '_seller_norm', '_sf_numeric']:
        if col in new_df.columns:
            new_df = new_df.drop(columns=[col])

    print(f"    Consolidated {consolidation_count} portfolio sale row(s) "
          f"({len(df)} -> {len(new_df)} rows)")

    if not dry_run:
        _rewrite_sheet(path, 'Major Sales', new_df)
    else:
        print(f"    [dry-run] Would consolidate {consolidation_count} row(s)")

    return consolidation_count


# ---------------------------------------------------------------------------
# Pipeline verification (Under Construction)
# ---------------------------------------------------------------------------

def _verify_pipeline_buildings(path, inventory_df, dry_run=False):
    """
    For Under Construction buildings, check if they exist in Supabase inventory.
    Prints warnings for:
    - Buildings not found in inventory
    - Buildings found but with aquila_competitive_set = False
    - Buildings found but with empty/null report_name
    This is informational only -- no data is modified.
    """
    if inventory_df.empty:
        print(f"    Skipping pipeline verification (no inventory data)")
        return

    try:
        xl = pd.ExcelFile(path)
    except Exception:
        return

    # Find the Under Construction sheet (first sheet that isn't proposed/planned)
    uc_sheet = None
    for sheet in xl.sheet_names:
        lower = sheet.lower()
        if 'proposed' not in lower and 'planned' not in lower and 'ignore' not in lower:
            uc_sheet = sheet
            break

    if not uc_sheet:
        return

    uc_df = pd.read_excel(path, sheet_name=uc_sheet)
    if uc_df.empty:
        return

    # Detect building name column (named format vs positional)
    col_names_lower = [str(c).lower() for c in uc_df.columns]
    has_named_cols = any('quarter' in c and 'delivery' in c for c in col_names_lower)

    if has_named_cols:
        name_col = next((c for c in uc_df.columns if str(c).lower() == 'name'), None)
    else:
        name_col = uc_df.columns[0] if len(uc_df.columns) > 0 else None

    if not name_col:
        return

    # Build normalized lookup set from inventory
    inv_names = set()
    for _, row in inventory_df.iterrows():
        pn = _normalize_name(row.get('property_name'))
        if pn:
            inv_names.add(pn)
        bp = _normalize_name(row.get('building_park'))
        if bp:
            inv_names.add(bp)

    # Check each UC building
    warnings = []
    for _, row in uc_df.iterrows():
        name_val = row.get(name_col)
        if pd.isna(name_val):
            continue
        name_str = str(name_val).strip()

        # Skip header-like rows (years, quarter markers)
        if re.match(r'^\d{4}$', name_str) or re.match(r'^\d[Qq]$', name_str):
            continue
        if not name_str:
            continue

        normalized = _normalize_name(name_str)
        if not normalized:
            continue

        # Find match in inventory
        if normalized not in inv_names:
            warnings.append(f"    [WARN] UC building not in inventory: '{name_str}'")
            continue

        # Check competitive set and report_name on matched rows
        matched = inventory_df[
            (inventory_df['property_name'].apply(_normalize_name) == normalized) |
            (inventory_df['building_park'].apply(_normalize_name) == normalized)
        ]
        if not matched.empty:
            row_inv = matched.iloc[0]
            if not row_inv.get('aquila_competitive_set', False):
                warnings.append(
                    f"    [WARN] UC building not in competitive set: '{name_str}'")
            rn = row_inv.get('report_name')
            if pd.isna(rn) or not str(rn).strip():
                warnings.append(
                    f"    [WARN] UC building missing report_name: '{name_str}'")

    if warnings:
        print(f"    Pipeline verification: {len(warnings)} warning(s)")
        for w in warnings:
            print(w)
    else:
        print(f"    Pipeline verification: all UC buildings found in inventory")


# ---------------------------------------------------------------------------
# Proposed/Planned sorting
# ---------------------------------------------------------------------------

def _sort_proposed(path, dry_run=False):
    """
    Sort the Proposed/Planned sheet by Submarket (A->Z), then by SF (ascending).
    Returns 1 if sorted, 0 if no change needed.
    """
    try:
        xl = pd.ExcelFile(path)
    except Exception:
        return 0

    # Find the Proposed/Planned sheet
    proposed_sheet = None
    for sheet in xl.sheet_names:
        lower = sheet.lower()
        if 'proposed' in lower or 'planned' in lower:
            proposed_sheet = sheet
            break

    if not proposed_sheet:
        return 0

    df = pd.read_excel(path, sheet_name=proposed_sheet)
    if df.empty or len(df) < 2:
        return 0

    # Detect columns (named or positional)
    sub_col = _find_column_by_keyword(df.columns, 'submarket')
    sf_col = _find_column_by_keyword(df.columns, 'size')
    if sf_col is None:
        sf_col = _find_sf_column(df.columns)

    # Fallback to positional if named detection failed
    if not sub_col and len(df.columns) >= 3:
        sub_col = df.columns[2]
    if not sf_col and len(df.columns) >= 2:
        sf_col = df.columns[1]

    if not sub_col or not sf_col:
        print(f"    Warning: Cannot identify Submarket/SF columns for Proposed sort")
        return 0

    # Filter out header-like rows
    skip_patterns = ['future developments', 'proposed/planned', 'proposed', 'planned',
                     'total proposed sf', 'total sf']
    mask = ~df.iloc[:, 0].astype(str).str.strip().str.lower().isin(skip_patterns)
    df_clean = df[mask].copy()

    if df_clean.empty or len(df_clean) < 2:
        return 0

    # Convert SF to numeric for sorting
    df_clean['_sf_sort'] = pd.to_numeric(
        df_clean[sf_col].astype(str).str.replace(',', '', regex=False),
        errors='coerce'
    )

    # Sort by Submarket ascending, then SF ascending
    df_sorted = df_clean.sort_values(
        [sub_col, '_sf_sort'], ascending=[True, True]
    ).drop(columns=['_sf_sort']).reset_index(drop=True)

    # Check if order actually changed
    df_clean_reset = df_clean.drop(columns=['_sf_sort']).reset_index(drop=True)
    if df_sorted.equals(df_clean_reset):
        return 0

    print(f"    Sorted Proposed/Planned by Submarket (A-Z), then SF (ascending)")

    if not dry_run:
        _rewrite_sheet(path, proposed_sheet, df_sorted)
    else:
        print(f"    [dry-run] Would re-sort {len(df_sorted)} Proposed/Planned rows")

    return 1


# ---------------------------------------------------------------------------
# Main cleanup orchestrator
# ---------------------------------------------------------------------------

def _get_files_to_process(cfg):
    """Return list of (label, path, file_role) for all quarterly Excel/CSV files.
    file_role: 'major_leases_sales', 'pipeline', or 'general'
    """
    report_type = _detect_report_type(cfg)

    files = [
        ('Major Leases & Sales', cfg.MAJOR_LEASES_SALES, 'major_leases_sales'),
    ]

    if report_type == 'office':
        files.append(('Office Availability', cfg.OFFICE_AVAIL, 'general'))
        files.append(('Building List', cfg.BUILDING_LIST, 'general'))
        files.append(('Citywide Pipeline', cfg.CITYWIDE_PIPELINE, 'pipeline'))
    else:  # industrial
        files.append(('Large Availabilities', cfg.LARGE_AVAIL, 'general'))
        files.append(('Building List', cfg.BUILDING_LIST, 'general'))
        files.append(('Development Pipeline', cfg.PIPELINE, 'pipeline'))

    # Quarterly Changes CSV files
    qc_dir = cfg.QUARTERLY_CHANGES_DIR
    if os.path.isdir(qc_dir):
        for fname in sorted(os.listdir(qc_dir)):
            if fname.lower().endswith('.csv'):
                files.append((fname, os.path.join(qc_dir, fname), 'general'))

    return files


def run_cleanup(cfg=None, dry_run=False):
    """
    Main entry point.
    cfg: report_config module (defaults to imported config).
    dry_run: if True, report changes without writing files.
    """
    if cfg is None:
        cfg = config

    report_type = _detect_report_type(cfg)
    label = '[DRY RUN] ' if dry_run else ''
    print('=' * 60)
    print(f'{label}QUARTERLY DATA CLEANUP -- {cfg.REPORT_LABEL}')
    print('=' * 60)

    # Load inventory from Supabase for building name matching
    print("\n  Loading inventory data from Supabase...")
    inventory_df = _load_inventory(report_type)
    inventory_lookup = _build_inventory_lookup(inventory_df)
    print(f"    Inventory lookup: {len(inventory_lookup)} building names indexed")

    files = _get_files_to_process(cfg)
    total_files = 0
    total_changes = 0

    for label_name, path, file_role in files:
        if not os.path.exists(path):
            print(f"  [SKIP] {label_name}: file not found")
            print(f"         {path}")
            continue

        print(f"\n  Processing: {label_name}")
        ext = os.path.splitext(path)[1].lower()

        if ext in ('.xlsx', '.xls'):
            n = process_excel_file(path, dry_run=dry_run)

            if file_role == 'major_leases_sales':
                _ensure_vertical_format(path, dry_run=dry_run)
                n += _cleanup_major_leases(path, inventory_lookup, dry_run=dry_run)
                n += _cleanup_major_sales(path, dry_run=dry_run)
            elif file_role == 'pipeline':
                _verify_pipeline_buildings(path, inventory_df, dry_run=dry_run)
                n += _sort_proposed(path, dry_run=dry_run)
        elif ext == '.csv':
            n = process_csv_file(path, dry_run=dry_run)
        else:
            print(f"    Unsupported file type: {ext}")
            continue

        total_files += 1
        total_changes += n
        if n == 0:
            print(f"    No changes needed.")

    print()
    print('=' * 60)
    action = 'Would update' if dry_run else 'Updated'
    print(f'CLEANUP COMPLETE -- {action} {total_changes} cell(s) across {total_files} file(s)')
    print('=' * 60)
    return total_changes


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Clean up quarterly report source data')
    parser.add_argument('--dry-run', action='store_true',
                        help='Preview changes without writing files')
    args = parser.parse_args()
    run_cleanup(dry_run=args.dry_run)
